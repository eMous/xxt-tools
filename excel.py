import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import numpy as np
from excel_utils import copy_cells, copy_sheet, delete_col_with_merged_ranges, delete_row_with_merged_ranges
import shutil
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import formatting, styles, Workbook
import warnings
warnings.simplefilter("ignore")

excel_dir = ""
# xlsname = f"学习通成绩 ({datetime.now().strftime("%Y_%m_%d %H_%M_%S")}).xlsx"
# xlsname = f"学习通成绩 (2024_01_20 16_05_10).xlsx"
estname = "成绩总表"


def get_xls_name():
    global xlsname
    xlsname = f"学习通成绩 ({datetime.now().strftime("%Y_%m_%d %H_%M_%S")}).xlsx"


def cr_cp_sheets_statistic(epath, path):
    if (os.path.exists(epath)):
        return
    # 创建空sheet
    ewb = Workbook()
    ews = ewb.active
    ews.title = estname

    wb = load_workbook(path)
    ws = wb['作业统计']

    copy_sheet(ws, ews)

    ews = ewb.create_sheet('作业统计')
    copy_sheet(ws, ews)

    ewb.save(epath)


def cr_cp_sheets_hwstyle(epath, paths):
    if (os.path.exists(epath)):
        return
    pd.DataFrame().to_excel(epath, sheet_name=estname, index=False)

    dfs_from = [pd.read_excel(path, sheet_name=0, engine="openpyxl")
                for path in paths]
    stnames_from = [list(pd.read_excel(path, sheet_name=None).keys())[
        0] for path in paths]

    df_e = pd.read_excel(epath)
    d_e = pd.read_excel(epath, sheet_name=None)
    ests = list(d_e.keys())

    return_dfs = sort_df(paths)

    # mydf (sheetname, date, df) from older date to newer date
    with pd.ExcelWriter(epath, engine="openpyxl", mode="a") as writer:
        for mydf in return_dfs:
            mydf[2].to_excel(writer, mydf[0], index=False)


def mindate(stname, df: pd.DataFrame):
    row_2_cells = df.iloc[0, :].tolist()

    date_header_cells = [(i, v) for i, v in enumerate(
        row_2_cells) if isinstance(v, str) and ("领取时间" in v)]
    if (len(date_header_cells) == 0):
        date_header_cells = [(i, v) for i, v in enumerate(
            row_2_cells) if isinstance(v, str) and ("时间" in v or "日期" in v)]

    min_dt = datetime.now()
    min_dt_from = ""
    for i, v in date_header_cells:
        date = df.iloc[:, i]
        for edate in date:
            try:
                dt = datetime.strptime(edate, '%Y-%m-%d %H:%M:%S')
                if (dt < min_dt):
                    min_dt = dt
                    min_dt_from = v
            except:
                pass
    print(f"作业{stname}的最早可能发布时间{min_dt},来自{min_dt_from}")
    return min_dt


def sort_df(paths):
    dfs = []
    for path in paths:
        xls = pd.ExcelFile(path)
        stname = xls.sheet_names[0]
        df = pd.read_excel(path, stname)
        date = mindate(stname, df)
        dfs.append((stname, date, df))

    returndfs = sorted(dfs, key=lambda x: x[1])
    return returndfs

# get valid paths of excel files


def get_paths(style: str):
    paths = []
    if (excel_dir == ""):
        return []
    if (not os.path.exists(excel_dir)):
        return []

    if (style == "hw"):
        paths = [entry for entry in os.scandir(excel_dir) if (entry.name.endswith(
            'xlsx') or entry.name.endswith('.xls')) and "学习通成绩" not in entry.name and ("一键导出" not in entry.name)]
    elif (style == "statistic"):
        paths = [entry for entry in os.scandir(excel_dir) if (entry.name.endswith(
            'xlsx') or entry.name.endswith('.xls')) and "学习通成绩" not in entry.name and ("一键导出" in entry.name)]
    else:
        raise RuntimeError("style must be hw or statistic")
    return paths


def get_epath(style):
    global excel_dir
    while True:
        paths = get_paths(style)
        if (len(paths) > 0):
            break
        else:
            excel_dir = input(f"当前查找excel的路径为{excel_dir}, 未找到导出数据，请重新设置路径:")
    epath = os.path.join(excel_dir, xlsname)
    return epath, paths


def create_summarysheet_statistic(epath: str):

    ewb = load_workbook(epath)
    st = ewb[estname]
    assert (st["C3"].value == "学校")
    delete_col_with_merged_ranges(st, column_index_from_string("C"))

    assert (st["C3"].value == "院系")
    delete_col_with_merged_ranges(st, column_index_from_string("C"))

    assert (st["C3"].value == "专业")
    delete_col_with_merged_ranges(st, column_index_from_string("C"))

    # print all cell value in row 4
    for cell in st[4]:
        # check cell has attribute col_idx
        if (not hasattr(cell, "col_idx")):
            continue
        if (cell.col_idx > 3):
            if (cell.value != "成绩"):
                delete_col_with_merged_ranges(st, cell.col_idx)

    delete_row_with_merged_ranges(st, 4)
    ewb.save(epath)


def complete_excel_task_statistic():
    get_xls_name()

    read_ini()
    epath, paths = get_epath("statistic")
    for path in paths:
        epath = epath[:epath.index(
            "学习通成绩")] + f"{path.name[:path.name.index("_统计一键导出")]}_" + epath[epath.index("学习通成绩"):]
        cr_cp_sheets_statistic(epath, path)
        create_summarysheet_statistic(epath)
        format_summarysheet(epath, "statistic")


def complete_excel_task_hwstyle():
    get_xls_name()

    read_ini()
    epath, paths = get_epath("hw")
    cr_cp_sheets_hwstyle(epath, paths)
    create_summarysheet_hwstyle(epath)
    format_summarysheet(epath, "hw")


def format_summarysheet_hwstyle(epath: str):
    wb = load_workbook(epath)
    ws = wb[estname]

    delete_row_with_merged_ranges(ws, 2)

    for column in ws.columns:
        if (not hasattr(column[5], "column_letter")):
            continue
        ws.column_dimensions[column[5].column_letter].width = 20

    align = Alignment(wrap_text=True, horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = align

    # set row 3 to 20 height
    ws.row_dimensions[3].height = 50
    for i in range(4, ws.max_row+1):
        ws.row_dimensions[i].height = 20

    # 冻结窗口
    ws.freeze_panes = ws['B2']

    # 补0
    minicell = 'D2'
    maxicell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
    score_cells = ws[minicell:maxicell]
    # iterate ws from range B4:S28
    for row in score_cells:
        for cell in row:
            if cell.value == None:
                cell.value = 0
            if cell.value == int(cell.value):
                cell.number_format = '0'
            else:
                cell.number_format = '0.00'

    # 条件格式
    red_color = 'ffc7ce'
    red_color_font = 'ff0000'
    red_font = styles.Font(color=red_color_font)
    red_fill = styles.PatternFill(
        start_color=red_color, end_color=red_color, fill_type='solid')
    rule = formatting.rule.CellIsRule(operator='lessThan', formula=[
                                      '60'], fill=red_fill)
    ws.conditional_formatting.add(f"{minicell}:{maxicell}", rule)
    # ws.conditional_formatting.add('B4:S28', formatting.rule.ColorScaleRule(start_type='num', start_value=0, start_color='00FF0000',
    #     end_type='num', end_value=60, end_color='00FFFFFF'
    # ))

    wb.save(epath)


def format_summarysheet(epath: str, style: str):

    wb = load_workbook(epath)
    ws = wb[estname]

    if style == "hw":
        delete_row_with_merged_ranges(ws, 2)

    for column in ws.columns:
        if (not hasattr(column[5], "column_letter")):
            continue
        ws.column_dimensions[column[5].column_letter].width = 17

    align = Alignment(wrap_text=True, horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = align

    # set row 3 to 20 height
    ws.row_dimensions[3].height = 50
    for i in range(4, ws.max_row+1):
        ws.row_dimensions[i].height = 20

    # 冻结窗口
    first_score_row = -1
    if style == "hw":
        first_score_row = 2
    elif style == "statistic":
        first_score_row = 4
    else:
        raise RuntimeError("style must be hw or statistic")

    ws.freeze_panes = ws[f"B{first_score_row}"]

    # 补0
    minicell = f"D{first_score_row}"
    maxicell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
    score_cells = ws[minicell:maxicell]
    # iterate ws from range B4:S28
    for row in score_cells:
        for cell in row:
            if cell.value == None:
                cell.value = 0
            if str(cell.value).isdigit():
                cell.number_format = '0'
            else:
                cell.number_format = '0.00'

    # 条件格式
    red_color = 'ffc7ce'
    red_color_font = 'ff0000'
    red_font = styles.Font(color=red_color_font)
    red_fill = styles.PatternFill(
        start_color=red_color, end_color=red_color, fill_type='solid')
    rule = formatting.rule.CellIsRule(operator='lessThan', formula=[
                                      '60'], fill=red_fill)
    ws.conditional_formatting.add(f"{minicell}:{maxicell}", rule)
    # ws.conditional_formatting.add('B4:S28', formatting.rule.ColorScaleRule(start_type='num', start_value=0, start_color='00FF0000',
    #     end_type='num', end_value=60, end_color='00FFFFFF'
    # ))

    wb.save(epath)


def format_summarysheet_statistic(epath: str):
    wb = load_workbook(epath)
    ws = wb[estname]

    for column in ws.columns:
        if (not hasattr(column[5], "column_letter")):
            continue
        ws.column_dimensions[column[5].column_letter].width = 17

    align = Alignment(wrap_text=True, horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = align

    # set row 3 to 20 height
    ws.row_dimensions[3].height = 50
    for i in range(4, ws.max_row+1):
        ws.row_dimensions[i].height = 20

    # 冻结窗口
    ws.freeze_panes = ws['B4']

    # 补0
    minicell = 'D4'
    maxicell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
    score_cells = ws[minicell:maxicell]
    # iterate ws from range B4:S28
    for row in score_cells:
        for cell in row:
            if cell.value == None:
                cell.value = 0
            if cell.value == int(cell.value):
                cell.number_format = '0'
            else:
                cell.number_format = '0.00'

    # 条件格式
    red_color = 'ffc7ce'
    red_color_font = 'ff0000'
    red_font = styles.Font(color=red_color_font)
    red_fill = styles.PatternFill(
        start_color=red_color, end_color=red_color, fill_type='solid')
    rule = formatting.rule.CellIsRule(operator='lessThan', formula=[
                                      '60'], fill=red_fill)
    ws.conditional_formatting.add(f"{minicell}:{maxicell}", rule)
    # ws.conditional_formatting.add('B4:S28', formatting.rule.ColorScaleRule(start_type='num', start_value=0, start_color='00FF0000',
    #     end_type='num', end_value=60, end_color='00FFFFFF'
    # ))

    wb.save(epath)


def create_summarysheet_hwstyle(epath: str):
    # 创建班级；学号；姓名 列 （从第一个sheet拷贝过来）
    stnames = pd.ExcelFile(epath).sheet_names
    assert (len(stnames) > 1)
    wb = load_workbook(epath)
    esheet = wb[estname]
    free_col = 1  # 未被使用的列序号
    df_from = pd.read_excel(epath, stnames[1])
    assert_names = ("学号", "姓名", "班级")
    cols_from = (0, 1, 5)
    for i in range(len(assert_names)):
        col_from = df_from.iloc[:, cols_from[i]]
        assert (assert_names[i] in col_from[0])
        for j, v in enumerate(col_from):
            # A1 A2 B1 B2 ...
            esheet[f"{get_column_letter(free_col)}{j+1}"] = v
        free_col += 1

    # 拷贝每一个Sheet中的作业成绩
    for stname in stnames[1:]:
        df = pd.read_excel(epath, stname)
        wkname = df.keys().tolist()[0]
        esheet[f"{get_column_letter(free_col)}1"] = wkname
        row_2_cells = df.iloc[0, :].tolist()
        score_header_cells = [(i, v) for i, v in enumerate(
            row_2_cells) if isinstance(v, str) and ("分数" in v)]
        assert (len(score_header_cells) > 0)
        score_header_cell = score_header_cells[0]
        assert (score_header_cell[0] == 8)
        for i, score in enumerate(df.iloc[:, 8]):
            # 第一行是 "分数" 已经被替换成了具体的作业名称
            if (i != 0):
                # assert 学号相同
                assert ((type(df.iloc[i, 0]) == float and np.isnan(df.iloc[i, 0]) and np.isnan(
                    esheet[f"A{i+1}"].value))
                    or
                    df.iloc[i, 0] == esheet[f"A{i+1}"].value)

                # assert 姓名相同
                assert ((type(df.iloc[i, 1]) == float and np.isnan(df.iloc[i, 1]) and np.isnan(
                    esheet[f"B{i+1}"].value))
                    or
                    df.iloc[i, 0] == esheet[f"A{i+1}"].value)

                esheet[f"{get_column_letter(free_col)}{i+1}"] = score
        free_col += 1
    wb.save(epath)


def read_ini():
    global excel_dir
    try:
        import configuration
        confs = {k: v for k, v in configuration.__dict__.items()
                 if not k.startswith("_")}
        for k, v in confs.items():
            globals()[k] = v
    except:
        pass


if __name__ == "__main__":
    # complete_excel_task_hwstyle()
    complete_excel_task_statistic()
